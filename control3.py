import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference

# Create workbook
wb = openpyxl.Workbook()

# Setup sheets
ws_dash = wb.active
ws_dash.title = "Dashboard"
ws_schedule = wb.create_sheet(title="Daily Schedule")
ws_settings = wb.create_sheet(title="Lists & Settings")

# Ensure grid lines are visible
ws_dash.views.sheetView[0].showGridLines = True
ws_schedule.views.sheetView[0].showGridLines = True
ws_settings.views.sheetView[0].showGridLines = True

# --- Theme Colors & Fonts ---
FONT_NAME = "Segoe UI"
COLOR_HEADER_BG = "2C3E50"      # Dark Slate/Navy
COLOR_HEADER_FG = "FFFFFF"      # White
COLOR_ACCENT = "16A085"         # Teal
COLOR_ZEBRA = "F8F9FA"          # Very light grey
COLOR_BORDER = "BDC3C7"         # Light grey border

font_title = Font(name=FONT_NAME, size=18, bold=True, color="2C3E50")
font_section = Font(name=FONT_NAME, size=13, bold=True, color="16A085")
font_header = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_HEADER_FG)
font_body = Font(name=FONT_NAME, size=11, color="000000")
font_body_bold = Font(name=FONT_NAME, size=11, bold=True, color="000000")
font_kpi_num = Font(name=FONT_NAME, size=20, bold=True, color="2C3E50")
font_kpi_lbl = Font(name=FONT_NAME, size=9, bold=False, color="7F8C8D")

fill_header = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
fill_accent = PatternFill(start_color=COLOR_ACCENT, end_color=COLOR_ACCENT, fill_type="solid")
fill_zebra = PatternFill(start_color=COLOR_ZEBRA, end_color=COLOR_ZEBRA, fill_type="solid")
fill_kpi = PatternFill(start_color="EAEDED", end_color="EAEDED", fill_type="solid")

thin_side = Side(border_style="thin", color=COLOR_BORDER)
border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_top_bottom = Border(top=thin_side, bottom=Side(border_style="double", color="2C3E50"))

# ==========================================
# 1. POPULATE SETTINGS & LISTS TAB
# ==========================================
ws_settings["A1"] = "Departments"
ws_settings["B1"] = "Status Options"
ws_settings["C1"] = "Priority Levels"

departments = ["Operations", "Finance", "Human Resources", "IT Support", "Sales & Marketing", "Customer Service"]
statuses = ["Not Started", "In Progress", "Completed", "On Hold"]
priorities = ["High", "Medium", "Low"]

for i, d in enumerate(departments, start=2): ws_settings[f"A{i}"] = d
for i, s in enumerate(statuses, start=2): ws_settings[f"B{i}"] = s
for i, p in enumerate(priorities, start=2): ws_settings[f"C{i}"] = p

for col in ["A", "B", "C"]:
    ws_settings[f"{col}1"].font = font_header
    ws_settings[f"{col}1"].fill = fill_header

# ==========================================
# 2. POPULATE DAILY SCHEDULE TAB
# ==========================================
headers_schedule = [
    "Task ID", "Date", "Department", "Task Description", 
    "Assigned To", "Start Time", "End Time", "Priority", "Status", "Comments"
]

for col_num, header in enumerate(headers_schedule, start=1):
    cell = ws_schedule.cell(row=3, column=col_num)
    cell.value = header
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Title
ws_schedule["A1"] = "DEPARTMENTAL DAILY TASK MASTER SCHEDULE"
ws_schedule["A1"].font = font_title

# Sample Data rows
sample_tasks = [
    ["TSK-001", "2026-06-23", "Operations", "Morning facility safety walk-through", "John Doe", "08:00", "09:00", "High", "Completed", "No issues found"],
    ["TSK-002", "2026-06-23", "Finance", "Reconcile previous day bank statements", "Jane Smith", "09:00", "11:00", "Medium", "In Progress", "Awaiting bank feed update"],
    ["TSK-003", "2026-06-23", "IT Support", "Server backup verification & patch update", "Alex Kim", "22:00", "23:30", "High", "Not Started", "Scheduled for off-peak hours"],
    ["TSK-004", "2026-06-23", "Human Resources", "Onboarding session for new cohort", "Sarah Jenkins", "10:00", "12:00", "Medium", "Completed", "3 new hires onboarded"],
    ["TSK-005", "2026-06-23", "Sales & Marketing", "Review Q3 social media ad metrics", "Michael Chang", "14:00", "15:30", "Low", "On Hold", "Waiting for budget finalization"],
    ["TSK-006", "2026-06-23", "Customer Service", "Resolve critical tier-3 support tickets", "Emily Ross", "11:00", "13:00", "High", "In Progress", "2/5 tickets resolved"],
]

for r_idx, row_data in enumerate(sample_tasks, start=4):
    for c_idx, val in enumerate(row_data, start=1):
        cell = ws_schedule.cell(row=r_idx, column=c_idx, value=val)
        cell.font = font_body
        cell.border = border_all
        if r_idx % 2 == 1:
            cell.fill = fill_zebra
        
        # Center align short text fields
        if c_idx in [1, 2, 6, 7, 8, 9]:
            cell.alignment = Alignment(horizontal="center", vertical="center")

# Add Data Validations
dv_dept = DataValidation(type="list", formula1="='Lists & Settings'!$A$2:$A$7", allow_blank=True)
dv_status = DataValidation(type="list", formula1="='Lists & Settings'!$B$2:$B$5", allow_blank=True)
dv_priority = DataValidation(type="list", formula1="='Lists & Settings'!$C$2:$C$4", allow_blank=True)

ws_schedule.add_data_validation(dv_dept)
ws_schedule.add_data_validation(dv_status)
ws_schedule.add_data_validation(dv_priority)

# Apply validation to 50 rows for scalability
dv_dept.add(f"C4:C50")
dv_priority.add(f"H4:H50")
dv_status.add(f"I4:I50")

# ==========================================
# 3. POPULATE DASHBOARD TAB
# ==========================================
ws_dash["A1"] = "DAILY TASK MANAGEMENT DASHBOARD"
ws_dash["A1"].font = font_title

# KPI Blocks
kpis = [
    ("Total Tasks", "=COUNTA('Daily Schedule'!A4:A50)", "B3", "B4"),
    ("Completed", "=COUNTIF('Daily Schedule'!I4:I50, \"Completed\")", "D3", "D4"),
    ("In Progress", "=COUNTIF('Daily Schedule'!I4:I50, \"In Progress\")", "F3", "F4"),
    ("Pending / On Hold", "=COUNTIF('Daily Schedule'!I4:I50, \"Not Started\") + COUNTIF('Daily Schedule'!I4:I50, \"On Hold\")", "H3", "H4")
]

for label, formula, col_lbl, col_val in kpis:
    # Label Cell
    lbl_cell = ws_dash[col_lbl]
    lbl_cell.value = label
    lbl_cell.font = font_kpi_lbl
    lbl_cell.alignment = Alignment(horizontal="center")
    lbl_cell.fill = fill_kpi
    
    # Value Cell
    val_cell = ws_dash[col_val]
    val_cell.value = formula
    val_cell.font = font_kpi_num
    val_cell.alignment = Alignment(horizontal="center")
    val_cell.fill = fill_kpi
    
    # Border
    lbl_cell.border = Border(left=thin_side, right=thin_side, top=thin_side)
    val_cell.border = Border(left=thin_side, right=thin_side, bottom=thin_side)

# Department Breakdowns Table
ws_dash["A7"] = "Departmental Task Summary"
ws_dash["A7"].font = font_section

dash_headers = ["Department", "Total Tasks", "Completed", "Pending"]
for c_idx, h in enumerate(dash_headers, start=1):
    cell = ws_dash.cell(row=8, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center")

for idx, dept in enumerate(departments, start=9):
    ws_dash.cell(row=idx, column=1, value=dept).font = font_body
    ws_dash.cell(row=idx, column=2, value=f"=COUNTIF('Daily Schedule'!$C$4:$C$50, A{idx})").font = font_body
    ws_dash.cell(row=idx, column=3, value=f"=COUNTIFS('Daily Schedule'!$C$4:$C$50, A{idx}, 'Daily Schedule'!$I$4:$I$50, \"Completed\")").font = font_body
    ws_dash.cell(row=idx, column=4, value=f"=B{idx}-C{idx}").font = font_body
    
    for c in range(1, 5):
        cell = ws_dash.cell(row=idx, column=c)
        cell.border = border_all
        if c > 1:
            cell.alignment = Alignment(horizontal="center")

# Total Row
tot_row = 9 + len(departments)
ws_dash.cell(row=tot_row, column=1, value="Total").font = font_body_bold
ws_dash.cell(row=tot_row, column=2, value=f"=SUM(B9:B{tot_row-1})").font = font_body_bold
ws_dash.cell(row=tot_row, column=3, value=f"=SUM(C9:C{tot_row-1})").font = font_body_bold
ws_dash.cell(row=tot_row, column=4, value=f"=SUM(D9:D{tot_row-1})").font = font_body_bold

for c in range(1, 5):
    cell = ws_dash.cell(row=tot_row, column=c)
    cell.border = border_top_bottom
    if c > 1:
        cell.alignment = Alignment(horizontal="center")

# Add Chart
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Tasks Status by Department"
chart.y_axis.title = "Number of Tasks"
chart.x_axis.title = "Department"

data = Reference(ws_dash, min_col=2, min_row=8, max_col=4, max_row=tot_row-1)
cats = Reference(ws_dash, min_col=1, min_row=9, max_row=tot_row-1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws_dash.add_chart(chart, "F7")
chart.width = 16
chart.height = 10

# ==========================================
# FINAL POLISH: COLUMN WIDTHS
# ==========================================
for ws in [ws_dash, ws_schedule, ws_settings]:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                # Basic length estimation avoiding formulas length expansion
                val_str = str(cell.value)
                if not val_str.startswith("="):
                    max_len = max(max_len, len(val_str))
                else:
                    max_len = max(max_len, 12)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

# Specific custom widths for optimal layout
ws_schedule.column_dimensions["D"].width = 35 # Description
ws_schedule.column_dimensions["J"].width = 30 # Comments
ws_dash.column_dimensions["A"].width = 22     # Department names

# Save workbook
filename = "Daily_Departmental_Task_Schedule.xlsx"
wb.save(filename)
print(f"Workbook saved successfully as {filename}")