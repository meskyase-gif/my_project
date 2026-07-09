import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

# Setup styles
font_family = "Segoe UI"
header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
title_font = Font(name=font_family, size=16, bold=True, color="1B365D")
section_font = Font(name=font_family, size=12, bold=True, color="1B365D")
bold_font = Font(name=font_family, size=10, bold=True)
regular_font = Font(name=font_family, size=10)
italic_font = Font(name=font_family, size=9, italic=True)

header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
sub_header_fill = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")
zebra_fill = PatternFill(start_color="F7F9FB", end_color="F7F9FB", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

thin_border_side = Side(border_style="thin", color="D9D9D9")
thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
thick_bottom = Border(bottom=Side(border_style="medium", color="1B365D"))
double_bottom = Border(top=Side(border_style="thin", color="D9D9D9"), bottom=Side(border_style="double", color="1B365D"))

# --- SHEET 1: INSTRUCTIONS & SETUP ---
ws_setup = wb.active
ws_setup.title = "Setup & Instructions"
ws_setup.views.sheetView[0].showGridLines = True

ws_setup["A1"] = "DAILY TASK MANAGEMENT & SCHEDULE SYSTEM"
ws_setup["A1"].font = title_font
ws_setup.row_dimensions[1].height = 30

ws_setup["A3"] = "How to Use This Template:"
ws_setup["A3"].font = section_font
instructions = [
    "1. Define your standard Departments and Team Members in the lists below (Columns A and B).",
    "2. Go to the 'Daily Task Master' tab to assign tasks, set priorities, pick departments, and assign owners.",
    "3. Use the 'Schedule View' tab to visualize tasks over the current week.",
    "4. Check the 'Dashboard' tab for automatic visual insights on completion rates and departmental workloads.",
    "Note: Dropdowns on subsequent sheets automatically pull from these setup lists."
]
for idx, inst in enumerate(instructions, start=4):
    ws_setup.cell(row=idx, column=1, value=inst).font = regular_font

# Lists for Data Validation
ws_setup["A11"] = "Departments"
ws_setup["A11"].font = bold_font
ws_setup["A11"].fill = sub_header_fill
ws_setup["B11"] = "Team Members"
ws_setup["B11"].font = bold_font
ws_setup["B11"].fill = sub_header_fill

depts = ["Operations", "Marketing", "Sales", "Finance", "HR", "IT Support", "Customer Success"]
members = ["Alice Smith", "Bob Jones", "Charlie Brown", "Diana Prince", "Evan Wright", "Fiona Gallagher", "George Clark"]

for i, d in enumerate(depts, start=12):
    ws_setup.cell(row=i, column=1, value=d).font = regular_font
for i, m in enumerate(members, start=12):
    ws_setup.cell(row=i, column=2, value=m).font = regular_font

# --- SHEET 2: TASK MASTER ---
ws_master = wb.create_sheet(title="Daily Task Master")
ws_master.views.sheetView[0].showGridLines = True

headers = ["Task ID", "Task Description", "Department", "Assigned To", "Date Assigned", "Due Date", "Priority", "Status", "Estimated Hours"]
ws_master.append(headers)
ws_master.row_dimensions[1].height = 24

for col_num, header in enumerate(headers, start=1):
    cell = ws_master.cell(row=1, column=col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

sample_tasks = [
    ["TSK-001", "Review quarterly budget reports", "Finance", "Evan Wright", "2026-06-22", "2026-06-23", "High", "In Progress", 4],
    ["TSK-002", "Deploy security patch to server", "IT Support", "George Clark", "2026-06-23", "2026-06-23", "Critical", "Not Started", 2],
    ["TSK-003", "Social media campaign scheduling", "Marketing", "Bob Jones", "2026-06-23", "2026-06-24", "Medium", "Completed", 3],
    ["TSK-004", "Follow up on enterprise leads", "Sales", "Diana Prince", "2026-06-23", "2026-06-25", "High", "In Progress", 5],
    ["TSK-005", "Conduct onboarding session", "HR", "Fiona Gallagher", "2026-06-24", "2026-06-24", "Low", "Not Started", 1.5],
    ["TSK-006", "Resolve high-priority ticket #402", "Customer Success", "Alice Smith", "2026-06-23", "2026-06-23", "High", "Completed", 1]
]

for row_data in sample_tasks:
    ws_master.append(row_data)

# Formats and Validations
dv_dept = DataValidation(type="list", formula1="'Setup & Instructions'!$A$12:$A$20", allow_blank=True)
dv_member = DataValidation(type="list", formula1="'Setup & Instructions'!$B$12:$B$20", allow_blank=True)
dv_priority = DataValidation(type="list", formula1='"Low,Medium,High,Critical"', allow_blank=True)
dv_status = DataValidation(type="list", formula1='"Not Started,In Progress,Completed,On Hold"', allow_blank=True)

ws_master.add_data_validation(dv_dept)
ws_master.add_data_validation(dv_member)
ws_master.add_data_validation(dv_priority)
ws_master.add_data_validation(dv_status)

dv_dept.add("C2:C100")
dv_member.add("D2:D100")
dv_priority.add("G2:G100")
dv_status.add("H2:H100")

# Format columns
for r in range(2, 50):
    ws_master.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    ws_master.cell(row=r, column=5).number_format = 'yyyy-mm-dd'
    ws_master.cell(row=r, column=6).number_format = 'yyyy-mm-dd'
    ws_master.cell(row=r, column=9).number_format = '#,##0.0'
    
    # Zebra striping
    fill = zebra_fill if r % 2 == 0 else white_fill
    for c in range(1, 10):
        cell = ws_master.cell(row=r, column=c)
        if cell.value is None and r > 7: continue
        cell.font = regular_font
        cell.border = thin_border
        if cell.fill.start_color.index == '00000000':  # Check if empty fill
            cell.fill = fill

# Conditional formatting for Status
red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

ws_master.conditional_formatting.add("H2:H50", CellIsRule(operator="equal", formula=['"Completed"'], fill=green_fill))
ws_master.conditional_formatting.add("H2:H50", CellIsRule(operator="equal", formula=['"In Progress"'], fill=yellow_fill))
ws_master.conditional_formatting.add("H2:H50", CellIsRule(operator="equal", formula=['"Not Started"'], fill=red_fill))

# --- SHEET 3: SCHEDULE VIEW ---
ws_sched = wb.create_sheet(title="Schedule View")
ws_sched.views.sheetView[0].showGridLines = True

ws_sched["A1"] = "WEEKLY SCHEDULE BY DEPARTMENT"
ws_sched["A1"].font = title_font

ws_sched["A3"] = "Department"
ws_sched["B3"] = "Mon (6/22)"
ws_sched["C3"] = "Tue (6/23)"
ws_sched["D3"] = "Wed (6/24)"
ws_sched["E3"] = "Thu (6/25)"
ws_sched["F3"] = "Fri (6/26)"

for col_num in range(1, 7):
    c = ws_sched.cell(row=3, column=col_num)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")

# Populate schedule with formulas mapping back to Master Task List
for idx, dept in enumerate(depts, start=4):
    ws_sched.cell(row=idx, column=1, value=dept).font = bold_font
    ws_sched.cell(row=idx, column=1).fill = sub_header_fill
    ws_sched.cell(row=idx, column=1).border = thin_border
    
    for col_idx, day_offset in enumerate(range(22, 27), start=2):
        date_str = f"2026-06-{day_offset}"
        # Simplified display formula to count tasks for that department on that day
        formula = f'=IF(COUNTIFS(\'Daily Task Master\'!$C$2:$C$50, $A{idx}, \'Daily Task Master\'!$F$2:$F$50, "2026-06-{day_offset}")>0, COUNTIFS(\'Daily Task Master\'!$C$2:$C$50, $A{idx}, \'Daily Task Master\'!$F$2:$F$50, "2026-06-{day_offset}") & " Task(s)", "-")'
        cell = ws_sched.cell(row=idx, column=col_idx, value=formula)
        cell.font = regular_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

# --- SHEET 4: DASHBOARD ---
ws_dash = wb.create_sheet(title="KPI Dashboard")
ws_dash.views.sheetView[0].showGridLines = True

ws_dash["A1"] = "TASK MANAGEMENT PERFORMANCE DASHBOARD"
ws_dash["A1"].font = title_font

# Summary Cards
ws_dash["A3"] = "Total Tasks"
ws_dash["A4"] = "=COUNTA('Daily Task Master'!$A$2:$A$50)"
ws_dash["B3"] = "Completed"
ws_dash["B4"] = "=COUNTIF('Daily Task Master'!$H$2:$H$50, \"Completed\")"
ws_dash["C3"] = "Completion Rate"
ws_dash["C4"] = "=B4/A4"

for col in ["A", "B", "C"]:
    ws_dash[f"{col}3"].font = italic_font
    ws_dash[f"{col}3"].alignment = Alignment(horizontal="center")
    ws_dash[f"{col}4"].font = Font(name=font_family, size=14, bold=True)
    ws_dash[f"{col}4"].alignment = Alignment(horizontal="center")
    ws_dash[f"{col}4"].border = thin_border
    ws_dash[f"{col}4"].fill = sub_header_fill

ws_dash["C4"].number_format = '0.0%'

# Department breakdown table
ws_dash["A7"] = "Department Breakdown"
ws_dash["A7"].font = section_font

ws_dash["A8"] = "Department"
ws_dash["B8"] = "Total Tasks Assigned"
ws_dash["C8"] = "Hours Allocated"
for col_num, header in enumerate(["Department", "Total Tasks Assigned", "Hours Allocated"], start=1):
    c = ws_dash.cell(row=8, column=col_num)
    c.font = header_font
    c.fill = header_fill

for idx, dept in enumerate(depts, start=9):
    ws_dash.cell(row=idx, column=1, value=dept).font = regular_font
    ws_dash.cell(row=idx, column=2, value=f'=COUNTIF(\'Daily Task Master\'!$C$2:$C$50, A{idx})').font = regular_font
    ws_dash.cell(row=idx, column=3, value=f'=SUMIF(\'Daily Task Master\'!$C$2:$C$50, A{idx}, \'Daily Task Master\'!$I$2:$I$50)').font = regular_font
    
    ws_dash.cell(row=idx, column=1).border = thin_border
    ws_dash.cell(row=idx, column=2).border = thin_border
    ws_dash.cell(row=idx, column=3).border = thin_border

# Auto-fit column widths across sheets
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                val_str = str(cell.value)
                if not val_str.startswith("="):
                    max_len = max(max_len, len(val_str))
                else:
                    max_len = max(max_len, 12) # default fallback for formulas
        sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

# Save workbook
file_path = "Department_Daily_Task_Schedule_System.xlsx"
wb.save(file_path)
print(f"File saved successfully as {file_path}")